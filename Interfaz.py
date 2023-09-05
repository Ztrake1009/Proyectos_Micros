import tkinter as tk
from tkinter import filedialog
import openpyxl
import time

class Grua:
    def __init__(self, ventana):
        self.ventana = ventana
        self.ventana.title("Grúa Pórtico")
        self.ventana.geometry("600x600")

        self.boton_Cargar = tk.Button(ventana, text="Cargar Archivo Excel", command=self.abrir_Excel)
        self.boton_Cargar.pack(side="top")

        self.letras_Suministro1 = []
        self.letras_Suministro2 = []
        self.letras_Carga = []

        self.boton_Modo1 = None
        self.boton_Modo2 = None

        self.boton_Volver = None
        self.boton_Iniciar = None

    def abrir_Excel(self):
        archivo_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if archivo_path:
            libro = openpyxl.load_workbook(archivo_path)
            hoja = libro.active
            self.letras_Carga = [cell.value for row in hoja.iter_rows() for cell in row]
            libro.close()
            self.Sel_Modo()
        #print (letras)

    def Sel_Modo(self):
        self.boton_Modo1 = tk.Button(ventana, text="Ejecutar Modo de Acomodo 1", command=self.Modo1)
        self.boton_Modo1.pack(side="top")
        self.boton_Modo2 = tk.Button(ventana, text="Ejecutar Modo de Acomodo 2", command=self.__init__)
        self.boton_Modo2.pack(side="top")

    def Modo1(self):
        self.boton_Cargar.destroy()
        self.boton_Modo1.destroy()
        self.boton_Modo2.destroy()

        self.boton_Volver = tk.Button(ventana, text="Volver a pantalla principal", command=self.volver_Inicio)
        self.boton_Volver.pack(side="top")

        self.boton_Iniciar = tk.Button(ventana, text="Iniciar Acomodo", command=self.mover_Cajas)
        self.boton_Iniciar.pack(side="top")

        self.tamX_Matriz1 = 8  # Cantidad columnas de la zona de suministro 1.
        self.tamY_Matriz1 = 3  # Cantidad filas de la zona de suministro 1.

        self.tamX_Matriz2 = 3  # Cantidad columnas de la zona de suministro 2.
        self.tamY_Matriz2 = 5  # Cantidad filas de la zona de suministro 2.

        self.tam_Matriz3 = 5  # Tamaño de la zona de carga.

        self.matriz1 = [[None for _ in range(self.tamX_Matriz1)] for _ in range(self.tamY_Matriz1)]  # Cantidad en X, Cantidad en Y.
        self.matriz2 = [[None for _ in range(self.tamX_Matriz2)] for _ in range(self.tamY_Matriz2)]
        self.matriz3 = [[None for _ in range(self.tam_Matriz3)] for _ in range(self.tam_Matriz3)]

        # Realmente cada espacio de la matriz de suministro mide 5 cm, en interfaz las medidas se multiplican por 8.
        self.tam_Celdas_Suministro = 40
        # Realmente cada espacio de la matriz de carga mide 5 cm, en interfaz las medidas se multiplican por 8.
        self.tam_Celdas_Carga = 40
        # Realmente el espacio libre es de 4 cm, en interfaz las medidas se multiplican por 8.
        self.tam_Esp_Libre = 32
        # Las bases de la grúa miden 8 cm cuadrados, en interfaz las medidas se multiplican por 8.
        self.tam_Bases = 64

        # Simula el tamaño de los 60x60 cm, igualmente cada lado se multiplica por 8.
        self.ancho = 480  # 60*8
        self.alto = 480  # 60*8
        self.lienzo = tk.Canvas(ventana, width=self.ancho, height=self.alto, background="black")
        self.lienzo.place(x=(600/2)-(self.ancho/2), y=(600/2)-(self.alto/2))
        #self.lienzo.place(x=1000-ancho, y=800-alto)

        # Crea un espacio libre en gris.
        # Las medidas son las escogidas según las especificaciones.
        x0 = self.tam_Bases
        y0 = self.tam_Bases
        x1 = 416  # (60*8)-(8*8)
        y1 = 416
        Esp_Libre = self.lienzo.create_rectangle(x0, y0, x1, y1, fill="#B2B2B2")

        # Listas para guardar las posiciones de cada espacio de las matrices.
        self.espacios_Matriz1 = []
        self.espacios_Matriz2 = []
        self.espacios_Matriz3 = []

        self.crear_Matriz_1()
        self.crear_Matriz_2()
        self.crear_Matriz_3()

        #print(self.espacios_Matriz1)
        #[116,116],[156,116]
        
        # Crea el objeto representativo de la grúa.
        self.radio = 10
        self.velocidad_X = 2
        self.velocidad_Y = 0
        self.X_Inicial = self.tam_Bases + (self.tam_Esp_Libre/2) - self.radio
        self.Y_Inicial = self.tam_Bases + (self.tam_Esp_Libre/2) - self.radio
        self.X_Final = self.tam_Bases + (self.tam_Esp_Libre/2) + self.radio
        self.Y_Final = self.tam_Bases + (self.tam_Esp_Libre/2) + self.radio
        self.grua = self.lienzo.create_oval(self.X_Inicial, self.Y_Inicial, self.X_Final, self.Y_Final, width=1, fill="red")

    # Carga el archivo de excel de prueba para la interfaz y modifica los colores.
    def leer_Excel_Prueba(self):
        archivo_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if archivo_path:
            libro = openpyxl.load_workbook(archivo_path)
            hoja = libro.active
            self.letras_Suministro1 = [cell.value for row in hoja.iter_rows() for cell in row]
            libro.close()

        self.letras_Suministro2.append(self.letras_Suministro1[24])
        self.letras_Suministro1.pop()
        i = 1
        while (i < 15):
            self.letras_Suministro2.append("")
            i += 1

        #print(self.letras_Suministro1)
        #print(self.letras_Suministro2)

    # Crea la Matriz 1.
    def crear_Matriz_1(self):
        self.leer_Excel_Prueba()
        letra_actual = 0

        for fila_Matriz1 in range(self.tamY_Matriz1):
            for col_Matriz1 in range(self.tamX_Matriz1):
                x0 = self.tam_Bases + self.tam_Esp_Libre + (col_Matriz1 * self.tam_Celdas_Suministro)
                y0 = self.tam_Bases + self.tam_Esp_Libre + (fila_Matriz1 * self.tam_Celdas_Suministro)
                x1 = x0 + self.tam_Celdas_Suministro
                y1 = y0 + self.tam_Celdas_Suministro
                color = self.ident_Color(letra_actual, self.letras_Suministro1)
                letra_actual += 1
                self.matriz1[fila_Matriz1][col_Matriz1] = self.lienzo.create_rectangle(x0, y0, x1, y1, fill=color)
                self.espacios_Matriz1.append([x0,y0])

    # Crea la Matriz 2.
    def crear_Matriz_2(self):
        letra_actual = 0

        for fila_Matriz2 in range(self.tamY_Matriz2):
            for col_Matriz2 in range(self.tamX_Matriz2):
                x0 = self.tam_Bases + self.tam_Esp_Libre + (col_Matriz2 * self.tam_Celdas_Suministro)
                y0 = self.tam_Bases + self.tam_Esp_Libre + (self.tam_Celdas_Suministro * self.tamY_Matriz1) + (fila_Matriz2 * self.tam_Celdas_Suministro)
                x1 = x0 + self.tam_Celdas_Suministro
                y1 = y0 + self.tam_Celdas_Suministro
                color = self.ident_Color(letra_actual, self.letras_Suministro2)
                letra_actual += 1
                self.matriz2[fila_Matriz2][col_Matriz2] = self.lienzo.create_rectangle(x0, y0, x1, y1, fill=color)
                self.espacios_Matriz2.append([x0,y0])

    # Crea la Matriz 3.
    def crear_Matriz_3(self):
        letra_actual = 0
        for fila_Matriz3 in range(self.tam_Matriz3):
            for col_Matriz3 in range(self.tam_Matriz3):
                x0 = self.tam_Bases + self.tam_Esp_Libre + (self.tam_Celdas_Suministro * self.tamX_Matriz2) + (col_Matriz3 * self.tam_Celdas_Carga)
                y0 = self.tam_Bases + self.tam_Esp_Libre + (self.tam_Celdas_Suministro * self.tamY_Matriz1) + (fila_Matriz3 * self.tam_Celdas_Carga)
                x1 = x0 + self.tam_Celdas_Carga
                y1 = y0 + self.tam_Celdas_Carga
                color = self.ident_Color(letra_actual, self.letras_Carga)
                letra_actual += 1
                self.matriz3[fila_Matriz3][col_Matriz3] = self.lienzo.create_rectangle(x0, y0, x1, y1, fill=color)
                self.espacios_Matriz3.append([x0,y0])
        
        #self.matriz3[0][0] = self.lienzo.create_rectangle(fill="red")

    def volver_Inicio(self):
        self.lienzo.destroy()
        self.boton_Volver.destroy()
        self.boton_Iniciar.destroy()

        self.letras_Suministro1 = []
        self.letras_Suministro2 = []
        self.letras_Carga = []

        self.boton_Cargar = tk.Button(ventana, text="Cargar Archivo Excel", command=self.abrir_Excel)
        self.boton_Cargar.pack(side="top")

    def ident_Color(self, letra_actual, letras):
        while letra_actual < len(letras):
            if letras[letra_actual] == "R":  # Rojo Claro.
                return "#FFAAAA"

            elif letras[letra_actual] == "G":  # Verde Claro.
                return "#AAFFAA"

            elif letras[letra_actual] == "B":  # Azul Claro.
                return "#AAAAFF"
            
            else:
                return "white"
            
    def mover_Cajas(self):
        for pos_Sum1, elem_Sum1 in enumerate(self.letras_Suministro1):
            for pos_Carga, elem_Carga in enumerate(self.letras_Carga):
                if (elem_Sum1 == elem_Carga):
                    x0 = self.espacios_Matriz1[pos_Sum1][0]
                    y0 = self.espacios_Matriz1[pos_Sum1][1]
                    self.ventana.after(500, self.actualizar_Color(x0,y0))
                    #time.sleep(0.01)

    def actualizar_Color(self, x0, y0):
        self.lienzo.create_rectangle(x0, y0, x0 + 40, y0 + 40, fill="white")







        """
        self.lienzo.move(self.grua, self.velocidad_X, self.velocidad_Y)
        self.X_Inicial += self.velocidad_X

        if self.X_Inicial + self.radio >= self.espacios_Matriz1[5][0] or self.X_Inicial - self.radio <= self.tam_Bases + (self.tam_Esp_Libre/2) - (2*self.radio):
            self.velocidad_X = -self.velocidad_X

        self.ventana.after(20, self.mover_Cajas)
        """

if __name__ == "__main__":
    ventana = tk.Tk()
    app = Grua(ventana)
    ventana.mainloop()