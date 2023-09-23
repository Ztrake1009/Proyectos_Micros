import tkinter as tk
from tkinter import filedialog
import openpyxl
import time

class Interfaz:
    def __init__(self, ventana):
        self.ventana = ventana
        self.ventana.title("Grúa Pórtico")
        self.ventana.geometry("600x600")

        self.boton_Cargar = tk.Button(ventana, text="Cargar Archivo Excel", command=self.abrir_Excel)
        self.boton_Cargar.pack(side="top")

        self.letras_Suministro1 = []
        self.letras_Suministro2 = []
        self.letras_Carga = []

        self.boton_Modo1 = None
        self.boton_Modo2 = None

        self.boton_Volver = None
        self.boton_Iniciar = None

    # Selecciona el archivo de excel para saber que cajas deben ir en la matriz de carga.
    def abrir_Excel(self):
        archivo_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if archivo_path:
            libro = openpyxl.load_workbook(archivo_path)
            hoja = libro.active
            self.letras_Carga = [cell.value for row in hoja.iter_rows() for cell in row]
            libro.close()
            self.Sel_Modo()
        #print (letras)

    def Sel_Modo(self):
        self.boton_Modo1 = tk.Button(ventana, text="Ejecutar Modo de Acomodo 1", command=self.Modo1)
        self.boton_Modo1.pack(side="top")
        self.boton_Modo2 = tk.Button(ventana, text="Ejecutar Modo de Acomodo 2", command=self.__init__)
        self.boton_Modo2.pack(side="top")

    def Modo1(self):
        self.boton_Cargar.destroy()
        self.boton_Modo1.destroy()
        self.boton_Modo2.destroy()

        self.Mensaje = tk.Label(ventana, text="En espera", font=("Arial", 16))
        self.Mensaje.pack(side="bottom")

        self.boton_Volver = tk.Button(ventana, text="Volver a pantalla principal", command=self.volver_Inicio)
        self.boton_Volver.pack(side="top")

        self.i = 0
        self.j = 0

        self.boton_Iniciar = tk.Button(ventana, text="Iniciar Acomodo", command=self.acomodar_Cajas)
        self.boton_Iniciar.pack(side="top")

        self.tamX_Matriz1 = 8  # Cantidad columnas de la zona de suministro 1.
        self.tamY_Matriz1 = 3  # Cantidad filas de la zona de suministro 1.

        self.tamX_Matriz2 = 3  # Cantidad columnas de la zona de suministro 2.
        self.tamY_Matriz2 = 5  # Cantidad filas de la zona de suministro 2.

        self.tam_Matriz3 = 5  # Tamaño de la zona de carga.

        self.matriz1 = [[None for _ in range(self.tamX_Matriz1)] for _ in range(self.tamY_Matriz1)]  # Cantidad en X, Cantidad en Y.
        self.matriz1.append([None]) #Crea un nuevo espacio en la matriz para que sean 25
        self.matriz2 = [[None for _ in range(self.tamX_Matriz2)] for _ in range(self.tamY_Matriz2)]
        self.matriz3 = [[None for _ in range(self.tam_Matriz3)] for _ in range(self.tam_Matriz3)]

        # Realmente cada espacio de la matriz de suministro mide 5 cm, en interfaz las medidas se multiplican por 8.
        self.tam_Celdas_Suministro = 40
        # Realmente cada espacio de la matriz de carga mide 5 cm, en interfaz las medidas se multiplican por 8.
        self.tam_Celdas_Carga = 40
        # Realmente el espacio libre es de 4 cm, en interfaz las medidas se multiplican por 8.
        self.tam_Esp_Libre = 32
        # Las bases de la grúa miden 8 cm cuadrados, en interfaz las medidas se multiplican por 8.
        self.tam_Bases = 64

        # Simula el tamaño de los 60x60 cm, igualmente cada lado se multiplica por 8.
        self.ancho = 480  # 60*8
        self.alto = 480  # 60*8
        self.lienzo = tk.Canvas(ventana, width=self.ancho, height=self.alto, background="black")
        self.lienzo.place(x=(600/2)-(self.ancho/2), y=(600/2)-(self.alto/2))
        #self.lienzo.place(x=1000-ancho, y=800-alto)

        # Crea un espacio libre en gris.
        # Las medidas son las escogidas según las especificaciones.
        x0 = self.tam_Bases
        y0 = self.tam_Bases
        x1 = 416  # (60*8)-(8*8)
        y1 = 416
        Esp_Libre = self.lienzo.create_rectangle(x0, y0, x1, y1, fill="#B2B2B2")

        # Listas para guardar las posiciones de cada espacio de las matrices.
        self.espacios_Matriz1 = []
        self.espacios_Matriz2 = []
        self.espacios_Matriz3 = []

        self.crear_Matriz_1()
        self.crear_Matriz_2()
        self.crear_Matriz_3()
        
        # Crea el objeto representativo de la grúa.
        self.movimiento_X = 1
        self.movimiento_Y = 1
        
        self.X_Actual = self.tam_Bases
        self.Y_Actual = self.tam_Bases
        self.Y_Final = self.X_Actual + self.tam_Celdas_Carga
        self.Y_Final = self.Y_Actual + self.tam_Celdas_Carga
        self.grua = self.lienzo.create_oval(self.X_Actual, self.Y_Actual, self.Y_Final, self.Y_Final, width=1, fill="orange")

    # Carga el archivo de excel de prueba para la matriz de suministro.
    def leer_Excel_Prueba(self):
        archivo_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if archivo_path:
            libro = openpyxl.load_workbook(archivo_path)
            hoja = libro.active
            self.letras_Suministro1 = [cell.value for row in hoja.iter_rows() for cell in row]
            libro.close()

        i = 0
        while (i < 16):
            self.letras_Suministro2.append("")
            i += 1

    # Crea la Matriz 1.
    def crear_Matriz_1(self):
        self.leer_Excel_Prueba()
        letra_actual = 0
        for fila_Matriz1 in range(self.tamY_Matriz1):
            for col_Matriz1 in range(self.tamX_Matriz1):
                x0 = self.tam_Bases + self.tam_Esp_Libre + (col_Matriz1 * self.tam_Celdas_Suministro)
                y0 = self.tam_Bases + self.tam_Esp_Libre + (fila_Matriz1 * self.tam_Celdas_Suministro)
                x1 = x0 + self.tam_Celdas_Suministro
                y1 = y0 + self.tam_Celdas_Suministro
                color = self.ident_Color(letra_actual, self.letras_Suministro1)
                letra_actual += 1
                self.matriz1[fila_Matriz1][col_Matriz1] = self.lienzo.create_rectangle(x0, y0, x1, y1, fill=color)
                self.espacios_Matriz1.append([x0,y0])

                # Crea el ultimo espacio de la matriz de manera que se "sale" de los 8x3 espacios originales, esto para que sean 25 espacios en total
                if (fila_Matriz1 == 2) and (col_Matriz1 == 7):
                    col_Matriz1 = 0
                    fila_Matriz1 = 3
                    x0 = self.tam_Bases + self.tam_Esp_Libre + (col_Matriz1 * self.tam_Celdas_Suministro)
                    y0 = self.tam_Bases + self.tam_Esp_Libre + (fila_Matriz1 * self.tam_Celdas_Suministro)
                    x1 = x0 + self.tam_Celdas_Suministro
                    y1 = y0 + self.tam_Celdas_Suministro                    
                    color = self.ident_Color(letra_actual, self.letras_Suministro1)
                    self.matriz1[fila_Matriz1][col_Matriz1] = self.lienzo.create_rectangle(x0, y0, x1, y1, fill=color)
                    self.espacios_Matriz1.append([x0,y0])
                
    # Crea la Matriz 2.
    def crear_Matriz_2(self):
        letra_actual = 0
        for fila_Matriz2 in range(self.tamY_Matriz2):
            for col_Matriz2 in range(self.tamX_Matriz2):
                # Se salta de crear el primer espacio, esto para que sea el ultimo espacio de la Matriz 1.
                if (fila_Matriz2 == 0) and (col_Matriz2 == 0):
                    col_Matriz2 = 1
                
                x0 = self.tam_Bases + self.tam_Esp_Libre + (col_Matriz2 * self.tam_Celdas_Suministro)
                y0 = self.tam_Bases + self.tam_Esp_Libre + (self.tam_Celdas_Suministro * self.tamY_Matriz1) + (fila_Matriz2 * self.tam_Celdas_Suministro)
                x1 = x0 + self.tam_Celdas_Suministro
                y1 = y0 + self.tam_Celdas_Suministro
                color = self.ident_Color(letra_actual, self.letras_Suministro2)
                letra_actual += 1
                self.matriz2[fila_Matriz2][col_Matriz2] = self.lienzo.create_rectangle(x0, y0, x1, y1, fill=color)
                self.espacios_Matriz2.append([x0,y0])

    # Crea la Matriz 3.
    def crear_Matriz_3(self):
        letra_actual = 0
        for fila_Matriz3 in range(self.tam_Matriz3):
            for col_Matriz3 in range(self.tam_Matriz3):
                x0 = self.tam_Bases + self.tam_Esp_Libre + (self.tam_Celdas_Suministro * self.tamX_Matriz2) + (col_Matriz3 * self.tam_Celdas_Carga)
                y0 = self.tam_Bases + self.tam_Esp_Libre + (self.tam_Celdas_Suministro * self.tamY_Matriz1) + (fila_Matriz3 * self.tam_Celdas_Carga)
                x1 = x0 + self.tam_Celdas_Carga
                y1 = y0 + self.tam_Celdas_Carga
                color = self.ident_Color(letra_actual, self.letras_Carga)
                letra_actual += 1
                self.matriz3[fila_Matriz3][col_Matriz3] = self.lienzo.create_rectangle(x0, y0, x1, y1, width=3, fill=color)
                self.espacios_Matriz3.append([x0,y0])

    # Vuelve al menu principal.
    def volver_Inicio(self):
        self.lienzo.destroy()
        self.boton_Volver.destroy()
        self.boton_Iniciar.destroy()
        self.Mensaje.destroy()

        self.letras_Suministro1 = []
        self.letras_Suministro2 = []
        self.letras_Carga = []

        self.boton_Cargar = tk.Button(ventana, text="Cargar Archivo Excel", command=self.abrir_Excel)
        self.boton_Cargar.pack(side="top")

    # Define el color del espacio de la matriz dependiendo de la lista que reciba.
    def ident_Color(self, letra_actual, letras):
        while letra_actual < len(letras):
            if letras[letra_actual] == "R":  # Rojo Claro.
                return "#FFAAAA"

            elif letras[letra_actual] == "G":  # Verde Claro.
                return "#AAFFAA"

            elif letras[letra_actual] == "B":  # Azul Claro.
                return "#AAAAFF"
            
            else:
                return "white"
            
    def acomodar_Cajas(self):
        self.boton_Iniciar.destroy()

        for pos_Sum1, elem_Sum1 in enumerate(self.letras_Suministro1):
            for pos_Carga, elem_Carga in enumerate(self.letras_Carga):
                if (elem_Sum1 == elem_Carga):
                    self.X_Destino = self.espacios_Matriz1[pos_Sum1][0]
                    self.Y_Destino = self.espacios_Matriz1[pos_Sum1][1]

                    # Ejecuta el movimiento, primero en X y luego en Y.
                    # Primero se mueve hacia la caja que debe recoger.
                    self.Mensaje.config(text="Moviendo Grua hacia la caja por recoger.")
                    self.mover_Grua_X()
                    time.sleep(0.1)
                    ventana.update()

                    self.mover_Grua_Y()
                    time.sleep(0.1)
                    self.Mensaje.config(text="Recogiendo caja.")
                    ventana.update()

                    self.actualizar_Suministro1()
                    time.sleep(0.1)
                    ventana.update()

                    self.X_Destino = self.espacios_Matriz3[pos_Carga][0]
                    self.Y_Destino = self.espacios_Matriz3[pos_Carga][1]
                    
                    # Ahora se mueve hacia el espacio donde debe dejar la caja.
                    self.Mensaje.config(text="Moviendo Grua hacia el espacio designado.")
                    self.mover_Grua_X()
                    time.sleep(0.1)
                    ventana.update()

                    self.mover_Grua_Y()
                    time.sleep(0.1)
                    self.Mensaje.config(text="Dejando caja.")
                    ventana.update()

                    self.actualizar_Carga(self.letras_Carga[pos_Carga])
                    time.sleep(0.1)
                    ventana.update()
                    self.letras_Carga[pos_Carga] = 0
                    break

        # Se mueve a la posición inicial.
        self.Mensaje.config(text="Volviendo a la posición inicial.")
        self.X_Destino = self.tam_Bases
        self.Y_Destino = self.tam_Bases
        self.mover_Grua_X()
        time.sleep(2)
        ventana.update()

        self.mover_Grua_Y()
        time.sleep(2)
        ventana.update()

    def actualizar_Suministro1(self):
        # Actualiza la matriz de la zona de Suministro 1.
        self.lienzo.create_rectangle(self.X_Destino, self.Y_Destino, self.X_Destino + 40, self.Y_Destino + 40, fill="white")
        
        self.grua = self.lienzo.create_oval(self.X_Actual, self.Y_Actual, self.X_Actual + 40, self.Y_Actual + 40, width=1, fill="orange")
        return
    
    def actualizar_Carga(self, letra):

        if letra == "R":  # Rojo Claro.
            color = "#FF0000"

        elif letra == "G":  # Verde Claro.
            color = "#00FF00"

        elif letra == "B":  # Azul Claro.
            color = "#0000FF"

        # Actualiza la matriz de la zona de Carga.
        self.lienzo.create_rectangle(self.X_Destino, self.Y_Destino, self.X_Destino + 40, self.Y_Destino + 40, width=3, fill=color)

        self.grua = self.lienzo.create_oval(self.X_Actual, self.Y_Actual, self.X_Actual + 40, self.Y_Actual + 40, width=1, fill="orange")

        return

    # Mover la grúa horizontalmente.
    def mover_Grua_X(self):
        # Movimiento horizontal.
        self.movimiento_X = self.X_Destino - self.X_Actual

        # Si ya está en la misma posición en X que el destino.
        if (self.X_Actual == self.X_Destino):
            self.movimiento_X = 0

        self.lienzo.move(self.grua, self.movimiento_X, 0)

        self.X_Actual = self.X_Destino
        return
    
    # Mover la grúa verticalmente.
    def mover_Grua_Y(self):
        # Movimiento vertical.
        self.movimiento_Y = self.Y_Destino - self.Y_Actual
        
        # Si ya está en la misma posición en Y que el destino.
        if (self.Y_Actual == self.Y_Destino):
            self.movimiento_Y = 0

        self.lienzo.move(self.grua, 0, self.movimiento_Y)

        self.Y_Actual = self.Y_Destino        
        return

    """
    # Mover la grúa en dirección Horizontal.
    def mover_Grua_X(self):

        # Movimiento hacia la derecha.
        if (self.X_Actual < self.X_Destino):
            self.velocidad_X = self.velocidad_X

        # Movimiento hacia la izquierda.
        if (self.X_Actual > self.X_Destino):
            self.velocidad_X = -self.velocidad_X
            
        if (self.X_Actual == self.X_Destino):
            self.ventana.after_cancel(self.mover_Grua_X)
            self.mover_Grua_Y()
            return
        
        self.lienzo.move(self.grua, self.velocidad_X, 0)

        self.X_Actual += self.velocidad_X

        self.ventana.after(20, self.mover_Grua_X)
        #self.mover_Grua_X()

    
    # Mover la grúa en dirección Vertical.
    def mover_Grua_Y(self):

        # Movimiento hacia la abajo.
        if (self.Y_Actual < self.Y_Destino):
            self.velocidad_Y = self.velocidad_Y

        # Movimiento hacia la arriba. 
        if (self.Y_Actual > self.Y_Destino):
            self.velocidad_Y = -self.velocidad_Y

        if (self.Y_Actual == self.Y_Destino):
            self.ventana.after_cancel(self.mover_Grua_X)
            return

        self.lienzo.move(self.grua, 0, self.velocidad_Y)

        self.Y_Actual += self.velocidad_Y

        self.ventana.after(20, self.mover_Grua_Y)
        #self.mover_Grua_Y()
    """

if __name__ == "__main__":
    ventana = tk.Tk()
    app = Interfaz(ventana)
    ventana.mainloop()