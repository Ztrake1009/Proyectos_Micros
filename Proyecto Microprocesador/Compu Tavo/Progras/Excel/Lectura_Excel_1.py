"""
import tkinter as tk
from tkinter import filedialog
import pandas as pd

def abrir_archivo_excel():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx *.xls")])
    if archivo:
        try:
            # Leer el archivo Excel usando pandas
            df = pd.read_excel(archivo)
            # Puedes hacer lo que quieras con el DataFrame 'df'
            # Por ejemplo, imprimirlo en la consola
            print(df)
        except Exception as e:
            # En caso de error al leer el archivo
            print("Error al leer el archivo:", e)

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Seleccionar archivo Excel")

# Botón para abrir el archivo
boton_abrir = tk.Button(ventana, text="Abrir archivo Excel", command=abrir_archivo_excel)
boton_abrir.pack(pady=20)

ventana.mainloop()


import tkinter as tk
from tkinter import filedialog
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

def abrir_archivo_excel():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx *.xls")])
    if archivo:
        try:
            # Leer el archivo Excel usando pandas
            df = pd.read_excel(archivo)
            # Mostrar los datos en una tabla
            texto_tabla.delete("1.0", tk.END)
            texto_tabla.insert(tk.END, df.to_string(index=False))

            # Mostrar los datos en un gráfico de barras
            plt.figure(figsize=(8, 4))
            plt.bar(df.columns, df.iloc[0])
            plt.xlabel('Columnas')
            plt.ylabel('Valores')
            plt.title('Gráfico de barras')
            plt.tight_layout()

            # Mostrar el gráfico en la interfaz gráfica
            canvas = FigureCanvasTkAgg(plt.gcf(), master=ventana)
            canvas.draw()
            canvas.get_tk_widget().pack()
        except Exception as e:
            # En caso de error al leer el archivo
            texto_tabla.delete("1.0", tk.END)
            texto_tabla.insert(tk.END, "Error al leer el archivo:\n" + str(e))

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Visualizar archivo Excel")

# Botón para abrir el archivo
boton_abrir = tk.Button(ventana, text="Abrir archivo Excel", command=abrir_archivo_excel)
boton_abrir.pack(pady=20)

# Área de texto para mostrar la tabla
texto_tabla = tk.Text(ventana, width=80, height=10)
texto_tabla.pack()

ventana.mainloop()


"""
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def abrir_archivo_excel():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx *.xls")])
    if archivo:
        try:
            # Leer el archivo Excel usando pandas y openpyxl
            global df, libro_excel, hoja
            df = pd.read_excel(archivo)
            libro_excel = openpyxl.load_workbook(archivo)
            hoja = libro_excel.active

            # Mostrar las celdas en la ventana de la interfaz
            mostrar_celdas()
        except Exception as e:
            # En caso de error al abrir o leer el archivo
            texto_celdas.delete("1.0", tk.END)
            texto_celdas.insert(tk.END, "Error al abrir el archivo:\n" + str(e))

def mostrar_celdas():
    # Limpiar el área de texto
    texto_celdas.delete("1.0", tk.END)

    # Obtener todas las celdas de la hoja en una lista de listas
    datos = [[celda.value for celda in fila] for fila in hoja.iter_rows(values_only=True)]

    # Mostrar las celdas en el área de texto de la interfaz gráfica
    texto_celdas.insert(tk.END, "\n".join([str(fila) for fila in datos]))

def guardar_cambios():
    # Obtener los datos del área de texto
    datos_actualizados = texto_celdas.get("1.0", tk.END)

    # Convertir los datos a un DataFrame de pandas
    nuevos_datos = [[valor.strip() for valor in fila.split("|")] for fila in datos_actualizados.split("\n")]
    nuevos_datos = [fila for fila in nuevos_datos if fila != [""]]  # Eliminar filas vacías

    # Actualizar los datos del DataFrame
    df_nuevos = pd.DataFrame(nuevos_datos[1:], columns=nuevos_datos[0])

    # Guardar los cambios en el archivo Excel
    for r_idx, row in enumerate(dataframe_to_rows(df_nuevos, index=False), 1):
        for c_idx, value in enumerate(row, 1):
            hoja.cell(row=r_idx, column=c_idx, value=value)

    libro_excel.save(datos_actualizados)

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Visualizar y Modificar archivo Excel")

# Botón para abrir el archivo
boton_abrir = tk.Button(ventana, text="Abrir archivo Excel", command=abrir_archivo_excel)
boton_abrir.pack(pady=10)

# Área de texto para mostrar y modificar las celdas
texto_celdas = tk.Text(ventana, width=100, height=20)
texto_celdas.pack()

# Botón para guardar cambios en el archivo
boton_guardar = tk.Button(ventana, text="Guardar cambios", command=guardar_cambios)
boton_guardar.pack(pady=10)

ventana.mainloop()

"""
from tkinter import *
from  tkinter import ttk


ws  = Tk()
ws.title('PythonGuides')
ws.geometry('500x500')
ws['bg'] = '#AC99F2'

game_frame = Frame(ws)
game_frame.pack()

my_game = ttk.Treeview(game_frame)

my_game['columns'] = ('player_id', 'player_name', 'player_Rank', 'player_states', 'player_city')

my_game.column("#0", width=0,  stretch=NO)
my_game.column("player_id",anchor=CENTER, width=80)
my_game.column("player_name",anchor=CENTER,width=80)
my_game.column("player_Rank",anchor=CENTER,width=80)
my_game.column("player_states",anchor=CENTER,width=80)
my_game.column("player_city",anchor=CENTER,width=80)

my_game.heading("#0",text="",anchor=CENTER)
my_game.heading("player_id",text="Id",anchor=CENTER)
my_game.heading("player_name",text="Name",anchor=CENTER)
my_game.heading("player_Rank",text="Rank",anchor=CENTER)
my_game.heading("player_states",text="States",anchor=CENTER)
my_game.heading("player_city",text="States",anchor=CENTER)

my_game.insert(parent='',index='end',iid=0,text='',
values=('1','Ninja','101','Oklahoma', 'Moore'))
my_game.insert(parent='',index='end',iid=1,text='',
values=('2','Ranger','102','Wisconsin', 'Green Bay'))
my_game.insert(parent='',index='end',iid=2,text='',
values=('3','Deamon','103', 'California', 'Placentia'))
my_game.insert(parent='',index='end',iid=3,text='',
values=('4','Dragon','104','New York' , 'White Plains'))
my_game.insert(parent='',index='end',iid=4,text='',
values=('5','CrissCross','105','California', 'San Diego'))
my_game.insert(parent='',index='end',iid=5,text='',
values=('6','ZaqueriBlack','106','Wisconsin' , 'TONY'))

my_game.pack()

ws.mainloop()
"""
