import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

class ExcelReaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Reader App")
        self.root.geometry("1000x800")

        self.excel_button = tk.Button(root, text="Seleccionar archivo Excel", command=self.load_excel)
        self.excel_button.pack()

        self.color_list = []
        self.color_matrix = None
        self.matrix_frame = None
        self.back_button = None

    def load_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            self.color_list = [cell.value for row in sheet.iter_rows() for cell in row]

            self.draw_matrix()

    def draw_matrix(self):
        self.excel_button.destroy()

        self.color_matrix = []
        for row in range(5):
            row_colors = self.color_list[row * 5: (row + 1) * 5]
            self.color_matrix.append(row_colors)

        self.matrix_frame = tk.Frame(root)
        self.matrix_frame.place(x=0, y=0)

        for i, row_colors in enumerate(self.color_matrix):
            for j, color in enumerate(row_colors):
                color_code = self.get_color_code(color)
                cell = tk.Canvas(self.matrix_frame, width=60, height=60, bg=color_code)
                cell.grid(row=i, column=j)
                #cell.place(x=1, y=1)

        self.back_button = tk.Button(root, text="Volver a la pantalla principal", command=self.back_to_main)
        self.back_button.pack()

    def back_to_main(self):
        self.matrix_frame.destroy()
        self.back_button.destroy()
        self.excel_button = tk.Button(root, text="Seleccionar archivo Excel", command=self.load_excel)
        self.excel_button.pack()
        self.color_matrix = None

    def get_color_code(self, color):
        if color == "R":
            return "red"
        elif color == "G":
            return "green"
        elif color == "B":
            return "blue"
        else:
            return "white"

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelReaderApp(root)
    root.mainloop()
